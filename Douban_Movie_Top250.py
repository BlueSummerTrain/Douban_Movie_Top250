#/usr/bin/env python
#coding:utf-8
#Function：使用requests、BeautifulSoup、openpyxl等库爬取豆瓣电影top250榜单并保存到本地Excel文档中py

import requests, bs4
from openpyxl import Workbook
from openpyxl.styles import colors, Font
 
def main():
    # 构建url
    url = "https://movie.douban.com/top250"
    # 向该页面发送get请求
    res = open_url(url)
    # 找到总共有多少页
    page_total = get_page(res)
    # 定义一个List存储最终所有电影的信息
    all_movies = []
    # 获取每一页的信息
    for i in range(page_total):
        # 观察每一页的url特点 拼凑出url
        url = 'https://movie.douban.com/top250?start=' + str(25 * i)
        # 向每个页面循环发送get请求
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.112 Safari/537.36'}
        res = requests.get(url, headers)
        # 将返回的页面信息传递给获取当前页电影信息的函数
        movies = get_movies(res)
        # 将每一页的电影信息添加到总的电影信息列表中
        all_movies.extend(movies)
    # 调用将电影数据写入到Excel文件中的方法
    print('电影信息已经爬取完毕，正在写入Excel文件中')
    save_to_excel(all_movies)
    print('电影信息已经成功保存到Excel文件中！')
 
 
# 打开链接请求数据的方法
def open_url(url):
    # 设置请求头 防止被服务器识别为爬虫 只需要设置User-Agent即可
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.112 Safari/537.36'}
    # 返回页面结果
    return requests.get(url, headers = headers).text
 
# 获取总页数的方法
def get_page(res):
    # 使用BeautifulSoup对返回的页面信息进行解析
    soup = bs4.BeautifulSoup(res, 'html.parser')  # 第一个参数为要解析的内容 第二个参数表示使用HTML解析器
    # 找到class为next的span元素
    next = soup.find('span', class_='next')
    # 它的前一个兄弟元素里的内容即为总页数（注意此处需要两个previous_sibling才能获取  因为元素内的空白字符被视作文本，而文本被视作节点。）
    total_page = next.previous_sibling.previous_sibling.text
    # 将总页数转换成整数返回
    return int(total_page)
 
# 获取某一页所有电影信息的方法
def get_movies(res):
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    # 获取当前页所有的电影名称
    movie_name = soup.find_all('div', class_='hd')
    # 获取当前页所有的电影评分
    movie_score = soup.find_all('span', class_='rating_num')
    # 获取当前页所有的电影简介
    movie_intro = soup.find_all('span', class_='inq')
    # 创建一个空的List用于保存电影信息
    movie = []
    for i in range(len(movie_name)):
        #  将每部电影的信息做成一个dict 放入movie中
        d = {
            'movie_name': movie_name[i].a.span.text,
            'movie_score' : movie_score[i].text,
            'movie_intro' : movie_intro[i].text
        }
        movie.append(d)
    return movie
# 将电影数据保存到Excel中的方法
def save_to_excel(all_movies):
    # 创建一个Excel工作簿
    wb = Workbook()
    # 创建一个sheet
    ws = wb.active
    # 设置sheet的标题为“豆瓣电影Top250排行榜”
    ws.title = u'豆瓣电影Top250排行榜'.encode('utf-8').decode("utf-8")
    # 设置sheet背景为红色
    ws.sheet_properties.tabColor = colors.RED
    # 将A1到D1合并单元格
    ws.merge_cells('A1:D1')
    # 设置A1的内容为“豆瓣电影Top250排行榜”
    ws['A1'] = '豆瓣电影Top250排行榜'
    # 创建一个字体格式 字体大小为18 字体颜色为红色
    font = Font(sz=18, color=colors.RED)
    # 给A1设置为该字体
    ws['A1'].font = font
    # 设置A1居中显示
 
    ws['A2'] = '排名'
    ws['B2'] = '电影名称'
    ws['C2'] = '电影评分'
    ws['D2'] = '电影简介'
 
    # 从第3行开始将电影数据逐行写入到Excel文件中
    # 定义起始位置
    start_pos = 3
    for (i, movie) in enumerate(all_movies):
        ws['A{0}'.format(start_pos + i)] = 'Top{0}'.format(i + 1)
        ws['B{0}'.format(start_pos + i)] = movie['movie_name']
        ws['C{0}'.format(start_pos + i)] = movie['movie_score']
        ws['D{0}'.format(start_pos + i)] = movie['movie_intro']
    # 保存工作簿
    wb.save('./豆瓣电影Top250排行.xlsx')
if __name__ == '__main__':
    main()
    #print()
